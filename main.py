import csv
import itertools
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.worksheet.worksheet import Worksheet

rus_dict_currency = {
    'AZN': {'name': 'Манаты', 'cost': 35.68},
    'BYR': {'name': 'Белорусские рубли', 'cost': 23.91},
    'EUR': {'name': 'Евро', 'cost': 59.90},
    'GEL': {'name': 'Грузинский лари', 'cost': 21.74},
    'KGS': {'name': 'Киргизский сом', 'cost': 0.76},
    'KZT': {'name': 'Тенге', 'cost': 0.13},
    'RUR': {'name': 'Рубли', 'cost': 1.00},
    'UAH': {'name': 'Гривны', 'cost': 1.64},
    'USD': {'name': 'Доллары', 'cost': 60.66},
    'UZS': {'name': 'Узбекский сум', 'cost': 0.0055}}


class Salary:
    def __init__(self, vacancies_dict):
        self.salary_from = vacancies_dict['salary_from']
        self.salary_to = vacancies_dict['salary_to']
        self.salary_currency = vacancies_dict['salary_currency']


class Vacancy:
    def __init__(self, vacancies_dict):
        self.name = vacancies_dict['name']
        self.salary = Salary(vacancies_dict)
        self.area_name = vacancies_dict['area_name']
        self.published_at = vacancies_dict['published_at']


class InputConnect:
    task_conditions = {
        'filename': {'prompt': 'Введите название файла', 'val': ''},
        'req_prof': {'prompt': 'Введите название профессии', 'val': ''}
    }

    @classmethod
    def get_task(self):
        for tc in self.task_conditions.keys():
            self.task_conditions[tc]['val'] = input(f'{self.task_conditions[tc]["prompt"]}: ')
        return True


class DataSet:
    @staticmethod
    def _filer(data):
        if len(data) == 0:
            return [{}]
        head = data.pop(0)
        return [dict(zip(head, row)) for row in data]

    @staticmethod
    def _csv_reader(filename):
        result = []
        with open(filename, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                if '' in row:
                    continue
                result.append(row)
        return result

    @staticmethod
    def _parser(task):
        res_data = DataSet._csv_reader(task.task_conditions['filename']['val'])
        dict_list = DataSet._filer(res_data)
        vac_list = []
        for d in dict_list:
            vac_list.append(Vacancy(d))
        return vac_list

    def __init__(self, task):
        self.file_name = task.task_conditions['filename']['val']
        self.vacancies_objects = DataSet._parser(task)
        self.dynamics_stats = DynamicStatistics(task, self.vacancies_objects)


class DynamicStatistics:
    @staticmethod
    def update_stats(sal_stat, vac_stat, main_ch, sal_m):
        if main_ch in sal_stat['val'].keys():
            sal_stat['val'][main_ch] += sal_m
            vac_stat['val'][main_ch] += 1
        else:
            sal_stat['val'][main_ch] = sal_m
            vac_stat['val'][main_ch] = 1

    @staticmethod
    def salary_normalizer(sal_stat, vac_stat, counted_stat):
        for k in counted_stat['val'].keys():
            sal_stat['val'][k] = int(sal_stat['val'][k] / (vac_stat['val'][k] * 2))
        sal_stat['val'] = dict(sorted(sal_stat['val'].items(), key=lambda x: x[0]))
        vac_stat['val'] = dict(sorted(vac_stat['val'].items(), key=lambda x: x[0]))

    def __init__(self, task, vacancies_objects):
        self.year_salary = {'name': 'Динамика уровня зарплат по годам', 'val': {}}
        self.year_vacancy = {'name': 'Динамика количества вакансий по годам', 'val': {}}
        self.selected_year_sal = {'name': 'Динамика уровня зарплат по годам для выбранной профессии', 'val': {}}
        self.selected_year_vac = {'name': 'Динамика количества вакансий по годам для выбранной профессии', 'val': {}}
        self.city_salary = {'name': 'Уровень зарплат по городам (в порядке убывания)', 'val': {}}
        self.city_vacancy = {'name': 'Доля вакансий по городам (в порядке убывания)', 'val': {}}
        for vac in vacancies_objects:
            year = int(vac.published_at[0:4])
            sal_m = ((float(vac.salary.salary_to) + float(vac.salary.salary_from)) *
                     rus_dict_currency[vac.salary.salary_currency]['cost'])
            self.update_stats(self.year_salary, self.year_vacancy, year, sal_m)
            if task.task_conditions['req_prof']['val'] in vac.name:
               self.update_stats(self.selected_year_sal, self.selected_year_vac, year, sal_m)
            city = vac.area_name
            self.update_stats(self.city_salary, self.city_vacancy, city, sal_m)
        self.salary_normalizer(self.year_salary, self.year_vacancy, self.year_salary)
        if len(self.selected_year_vac['val']) == 0:
            self.selected_year_sal['val'] = {2022: 0}
            self.selected_year_vac['val'] = {2022: 0}
        else:
            self.salary_normalizer(self.selected_year_sal, self.selected_year_vac, self.selected_year_vac)
        for c in self.city_salary['val'].keys():
            self.city_salary['val'][c] = int(self.city_salary['val'][c] / (self.city_vacancy['val'][c] * 2))
        self.city_vacancy['val'] = dict(
            filter(lambda x: x[1] >= len(vacancies_objects) / 100, self.city_vacancy['val'].items()))
        self.city_salary['val'] = dict(
            filter(lambda x: self.city_vacancy['val'].__contains__(x[0]), self.city_salary['val'].items()))
        self.city_salary['val'] = dict(sorted(self.city_salary['val'].items(), key=lambda x: x[1], reverse=True))
        self.city_vacancy['val'] = dict(sorted(self.city_vacancy['val'].items(), key=lambda x: (-x[1])))
        for c in self.city_vacancy['val']:
            self.city_vacancy['val'][c] = round(self.city_vacancy['val'][c] / len(vacancies_objects), 4)
        self.city_salary['val'] = dict(itertools.islice(self.city_salary['val'].items(), 10))
        self.city_vacancy['val'] = dict(itertools.islice(self.city_vacancy['val'].items(), 10))


class Report:
    ws1: Worksheet
    ws2: Worksheet
    def __init__(self, dataset):
        self.year_salary = dataset.dynamics_stats.year_salary['val']
        self.year_vacancy = dataset.dynamics_stats.year_vacancy['val']
        self.selected_year_sal = dataset.dynamics_stats.selected_year_sal['val']
        self.selected_year_vac = dataset.dynamics_stats.selected_year_vac['val']
        self.city_salary = dataset.dynamics_stats.city_salary['val']
        self.city_vacancy = dataset.dynamics_stats.city_vacancy['val']

    def generate_excel(self, req_prof):
        workbook = Workbook()
        statistics_by_year = workbook.worksheets[0]
        statistics_by_year.title = "Cтатистика по годам"
        statistics_by_city = workbook.create_sheet("Cтатистика по городам")
        statistics_by_year.append(["Год", "Средняя зарплата", "Количество вакансий", f"Средняя зарплата - {req_prof}", f"Количество вакансий - {req_prof}"])
        for i, year in enumerate(self.year_salary.keys(), 2):
            statistics_by_year.cell(row=i, column=1, value=year)
            for j, dict in enumerate((self.year_salary, self.year_vacancy, self.selected_year_sal, self.selected_year_vac), 2):
                statistics_by_year.cell(row=i, column=j, value=dict[year])
        statistics_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.city_salary.keys(), 2):
            statistics_by_city.cell(row=i, column=1, value=city)
            statistics_by_city.cell(row=i, column=2, value=self.city_salary[city])
        for i, city in enumerate(self.city_vacancy.keys(), 2):
            statistics_by_city.cell(row=i, column=4, value=city)
            statistics_by_city.cell(row=i, column=5, value=self.city_vacancy[city]).number_format = '0.00%'
        self.table_styler(workbook)
        self.ws1 = statistics_by_year
        self.ws2 = statistics_by_city
        workbook.save('report.xlsx')

    @staticmethod
    def table_styler(wb):
        bold_font = Font(bold=True)
        thin_border_style = Side(border_style="thin", color="000000")
        outline = Border(top=thin_border_style, left=thin_border_style, right=thin_border_style, bottom=thin_border_style)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, profession):
        fig1, ((gr11, gr12), (gr21, gr22)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self.generate_year_salary_graph(gr11, profession)
        self.generate_year_vacancy_graph(gr12, profession)
        self.generate_city_salary_graph(gr21)
        self.generate_city_vacancy_graph(gr22)
        plt.savefig('graph.png')

    def generate_year_salary_graph(self, gr, profession):
        gr_labels = self.year_salary.keys()
        x = np.arange(len(gr_labels))
        width = 0.35
        gr.bar(x - width / 2, self.year_salary.values(), width, label='Средняя з/п')
        gr.bar(x + width / 2, self.selected_year_sal.values(), width, label=f'З/п {profession}')
        gr.set_xticks(x, gr_labels, fontsize=8, rotation=90, ha='right')
        gr.legend(fontsize=8, loc='upper left')
        gr.yaxis.grid(True)
        gr.set_title("Уровень зарплат по годам")

    def generate_year_vacancy_graph(self, gr, profession):
        gr_labels = self.year_vacancy.keys()
        x = np.arange(len(gr_labels))
        width = 0.35
        gr.bar(x - width / 2, self.year_vacancy.values(), width, label='Количество вакансий')
        gr.bar(x + width / 2, self.selected_year_vac.values(), label=f'Количество вакансий {profession}')
        gr.set_xticks(x, gr_labels, fontsize=8, rotation=90, ha='right')
        gr.legend(fontsize=8, loc='upper left')
        gr.yaxis.grid(True)
        gr.set_title("Количество вакансий по годам")

    def generate_city_salary_graph(self, gr):
        gr_labels = self.city_salary.keys()
        y_pos = np.arange(len(gr_labels))
        gr.barh(y_pos, self.city_salary.values(), align='center')
        gr.set_yticks(y_pos, fontsize=6, labels=gr_labels)
        gr.invert_yaxis()
        gr.xaxis.grid(True)
        gr.set_title("Уровень зарплат по городам")

    def generate_city_vacancy_graph(self, gr):
        f_labels = list(self.city_vacancy.keys())
        values = list(self.city_vacancy.values())
        f_labels.append('Другие')
        values.append(1 - sum(values))
        gr.pie(values, labels=f_labels, textprops={'fontsize': 6}, startangle=0, labeldistance=1.1,
               colors=['tab:orange', 'tab:green', 'tab:red', 'tab:purple', 'tab:brown', 'tab:pink',
                      'tab:gray', 'tab:olive', 'tab:cyan', 'tab:blue', 'tab:blue'])
        gr.set_title("Доля вакансий по городам")

    def remake_to_percantage(self, ws, column_index):
        for row in range(2, ws.max_row + 1):
            ws.cell(column=column_index, row=row).value = str(round((ws.cell(column=column_index, row=row).value * 100), 2)).replace(".", ",") + "%"
        return ws

    def generate_pdf(self, profession):
        year_stat = self.ws1
        cities_stat = self.ws2
        cities_stat = self.remake_to_percantage(cities_stat, 5)
        image = "graph.png"
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(profession=profession, image=image, year_stat=year_stat,
                                       cities_stat=cities_stat)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\ав\Desktop\программы\Python\pythonProject\Addons\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config,
                           options={'enable-local-file-access': None})
        # config = pdfkit.configuration(wkhtmltopdf=r'C:\Users\ав\Desktop\программы\Python\pythonProject\Addons\wkhtmltopdf\bin\wkhtmltopdf.exe')



my_task = InputConnect()
my_task.get_task()
my_data = DataSet(my_task)
if (my_data.vacancies_objects == None): exit()
print(f'{my_data.dynamics_stats.year_salary["name"]}: {my_data.dynamics_stats.year_salary["val"]}')
print(f'{my_data.dynamics_stats.year_vacancy["name"]}: {my_data.dynamics_stats.year_vacancy["val"]}')
print(f'{my_data.dynamics_stats.selected_year_sal["name"]}: {my_data.dynamics_stats.selected_year_sal["val"]}')
print(f'{my_data.dynamics_stats.selected_year_vac["name"]}: {my_data.dynamics_stats.selected_year_vac["val"]}')
print(f'{my_data.dynamics_stats.city_salary["name"]}: {my_data.dynamics_stats.city_salary["val"]}')
print(f'{my_data.dynamics_stats.city_vacancy["name"]}: {my_data.dynamics_stats.city_vacancy["val"]}')
report = Report(my_data)
main_input = input('Введите тип вывода: ')
if main_input != "Вакансии" and main_input != "Cтатистика":
    print("Введён неправильный тип вода")
if main_input == "Вакансии":
    report.generate_excel(my_task.task_conditions['req_prof']['val'])
else:
    report.generate_image(my_task.task_conditions['req_prof']['val'])

